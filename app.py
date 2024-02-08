from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL
from openpyxl import load_workbook, Workbook # read excel
from configExcel import Config


app = Flask(__name__)

# Mysql connection
# Usar variables de entorno (.env)
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = '123987'
app.config['MYSQL_DB'] = 'reunion_general'

mysql = MySQL(app)

# settings
app.secret_key = 'mysecretkey' # para habilitar el mensaje de flash()

@app.route('/')
def index():
    cur = mysql.connection.cursor()
    sql = """SELECT * FROM reunion_general.hermanos;"""
    cur.execute(sql)
    data = cur.fetchall()

    totalHermanos = len(data)
    return render_template('index.html', usuarios=data, totalUsuarios=totalHermanos)


@app.route('/agregar_usuario', methods = ['POST'])
def add_user():
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        obervacion = request.form['observacion']

        cur = mysql.connection.cursor()
        sql = """
                INSERT INTO reunion_general.hermanos (nombre, apellido, observacion)
                VALUES (%s, %s, %s)
            """
        params = (nombre, apellido, obervacion)
        cur.execute(sql, params)
        mysql.connection.commit() # Save changes
        flash('Usuario agregado exitosamente!')
    return redirect(url_for('index')) # redireccionar a la funcion principal 'index'.


@app.route('/edit/<id>')
def get_user(id):
    cur = mysql.connection.cursor()
    sql = """
            SELECT * FROM reunion_general.hermanos
            WHERE hermano_id = %s
        """
    params = id,
    cur.execute(sql, params)
    data = cur.fetchall()
    return render_template('edit_contact.html', usuario = data[0])

@app.route('/update/<id>', methods = ['POST'])
def update_contact(id):
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        observacion = request.form['observacion']

        cur = mysql.connection.cursor()
        sql = """
                UPDATE reunion_general.hermanos
                SET nombre = %s, apellido = %s, observacion = %s
                WHERE hermano_id = %s
            """
        params = (nombre, apellido, observacion, id)
        cur.execute(sql, params)
        mysql.connection.commit() # Save changes
        flash('Usuario actualizado exitosamente!')
    return redirect(url_for('index')) # Redireccionamos a la pantalla principal

@app.route('/delete/<string:id>')
def delete_contact(id):
    cur = mysql.connection.cursor()
    sql = """
            DELETE FROM reunion_general.hermanos
            WHERE hermano_id = %s
        """
    params = id,
    cur.execute(sql, params)
    mysql.connection.commit() # Save changes
    flash('Usuario Eliminado exitosamente!')
    return redirect(url_for('index'))

@app.route('/vaciarLista')
def vaciar_lista():
    cur = mysql.connection.cursor()
    sql = """DELETE FROM reunion_general.hermanos;"""
    cur.execute(sql)
    mysql.connection.commit() # Save changes
    flash('Lista limpia.')
    return redirect(url_for('index'))

@app.route('/exportar')
def exportar_a_excel():
    cur = mysql.connection.cursor()
    sql = """SELECT * FROM reunion_general.hermanos"""
    cur.execute(sql)
    data = cur.fetchall()

    # Abre el archivo Excel existente o crea uno nuevo si no existe
    file_name = Config.NOMBRE_ARCHIVO
    hojaExcel = Config.HOJA

    libro_excel = load_workbook(file_name[0])
    hoja = libro_excel.create_sheet(hojaExcel)

    # ENCABEZADOS
    hoja['A1'] = "Nombre"
    hoja['B1'] = "Apellido"
    hoja['C1'] = "Ingreso"
    hoja['D1'] = "Observacion"

    # Añadir los nuevos datos debajo de los existentes
    for hermano in data:
        hoja.append([hermano[1], hermano[2], hermano[3], hermano[4]])

    # Guardar el libro de Excel
    libro_excel.save(file_name[0])

    flash('Usuarios exportados exitosamente!')
    return redirect(url_for('index'))

if __name__ == "__main__":
    app.run(port = 4000, debug = True)